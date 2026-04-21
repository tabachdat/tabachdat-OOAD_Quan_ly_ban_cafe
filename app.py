from __future__ import annotations

import hashlib
import io
from datetime import date, datetime, timedelta
from decimal import Decimal
from functools import wraps
from pathlib import Path

from flask import (
    Flask,
    abort,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import bindparam, text
from werkzeug.security import check_password_hash, generate_password_hash

from config import Config
from models.db import db

app = Flask(__name__)
app.config.from_object(Config)
db.init_app(app)

ROLE_ADMIN = "Admin"
ROLE_STAFF = "Staff"

TABLE_EMPTY = "Trống"
TABLE_OCCUPIED = "Đang dùng"

ORDER_PENDING = "Chờ xác nhận"
ORDER_PROCESSING = "Đang chế biến"
ORDER_COMPLETED = "Đã hoàn thành"
ORDER_PAID = "Đã thanh toán"
ORDER_CANCELLED = "Đã hủy"

NOTIFY_ORDER = "ORDER"
NOTIFY_SUPPORT = "SUPPORT"
NOTIFY_PAYMENT = "PAYMENT"

DISCOUNT_CODES = {
    "GIAM10": Decimal("0.10"),
    "SALE20": Decimal("0.20"),
    "VIP15": Decimal("0.15"),
}

SUGAR_OPTIONS = [
    "Bình thường",
    "Ít đường",
    "Không đường",
]

ICE_OPTIONS = [
    "Bình thường",
    "Ít đá",
    "Không đá",
]

ORDER_STATUS_ALIASES = {
    "Đang chờ": ORDER_PENDING,
    "Pending": ORDER_PENDING,
    "Processing": ORDER_PROCESSING,
    "Đang pha chế": ORDER_PROCESSING,
    "Hoàn thành": ORDER_COMPLETED,
    "Completed": ORDER_COMPLETED,
    "Đã phục vụ": ORDER_COMPLETED,
    "Paid": ORDER_PAID,
    "Cancelled": ORDER_CANCELLED,
}

TABLE_STATUS_ALIASES = {
    "Available": TABLE_EMPTY,
    "Occupied": TABLE_OCCUPIED,
    "Có khách": TABLE_OCCUPIED,
}


@app.template_filter("currency")
def currency_filter(value: Decimal | int | float | None) -> str:
    amount = Decimal(value or 0)
    return f"{amount:,.0f} đ"


@app.template_filter("datetime_vn")
def datetime_filter(value: datetime | None, pattern: str = "%d/%m/%Y %H:%M") -> str:
    if not value:
        return "--"
    return value.strftime(pattern)


@app.template_filter("date_vn")
def date_filter(value: date | datetime | None, pattern: str = "%d/%m/%Y") -> str:
    if not value:
        return "--"
    if isinstance(value, datetime):
        return value.strftime(pattern)
    return value.strftime(pattern)


def query_all(sql: str, params: dict | None = None):
    with db.engine.connect() as connection:
        return connection.execute(text(sql), params or {}).mappings().all()


def query_one(sql: str, params: dict | None = None):
    with db.engine.connect() as connection:
        return connection.execute(text(sql), params or {}).mappings().first()


def query_scalar(sql: str, params: dict | None = None):
    with db.engine.connect() as connection:
        return connection.execute(text(sql), params or {}).scalar()


def require_role(*allowed_roles: str):
    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            role = session.get("role")
            if not role or (allowed_roles and role not in allowed_roles):
                flash("Vui lòng đăng nhập bằng tài khoản hợp lệ.", "error")
                return redirect(url_for("login"))
            return view(*args, **kwargs)

        return wrapped

    return decorator


def normalize_order_status(status: str | None) -> str:
    if not status:
        return ORDER_PENDING
    return ORDER_STATUS_ALIASES.get(status, status)


def normalize_table_status(status: str | None) -> str:
    if not status:
        return TABLE_EMPTY
    return TABLE_STATUS_ALIASES.get(status, status)


def enrich_order(row) -> dict:
    item = dict(row)
    item["Status"] = normalize_order_status(item.get("Status"))
    return item


def enrich_orders(rows) -> list[dict]:
    return [enrich_order(row) for row in rows]


def enrich_table(row) -> dict:
    item = dict(row)
    item["Status"] = normalize_table_status(item.get("Status"))
    item["OrderStatus"] = normalize_order_status(item.get("OrderStatus"))
    return item


def check_password(stored_password: str, plain_password: str) -> bool:
    if not stored_password:
        return False

    if stored_password.startswith("pbkdf2:") or stored_password.startswith("scrypt:"):
        return check_password_hash(stored_password, plain_password)

    return stored_password == plain_password


def hash_password(password: str) -> str:
    return generate_password_hash(password)


def get_product_images() -> list[str]:
    image_dir = Path(app.root_path) / "static" / "images" / "products"
    return sorted(file.name for file in image_dir.iterdir() if file.is_file())


def make_nav_links(role: str | None) -> list[dict]:
    if role == ROLE_ADMIN:
        return [
            {"endpoint": "admin_dashboard", "label": "Tổng quan"},
            {"endpoint": "admin_products", "label": "Quản lý thực đơn"},
            {"endpoint": "admin_users", "label": "Tài khoản nhân viên"},
            {"endpoint": "admin_report", "label": "Báo cáo doanh thu"},
        ]

    if role == ROLE_STAFF:
        return [
            {"endpoint": "staff_dashboard", "label": "Đơn hàng và bàn"},
        ]

    return []


def get_table_or_404(table_id: int) -> dict:
    table = query_one(
        """
        SELECT TableID, TableName, Status, QRCodePath
        FROM CafeTables
        WHERE TableID = :table_id
        """,
        {"table_id": table_id},
    )

    if not table:
        abort(404)

    item = dict(table)
    item["Status"] = normalize_table_status(item.get("Status"))
    return item


def get_order_or_404(order_id: int) -> dict:
    order = query_one(
        """
        SELECT o.*, t.TableName
        FROM Orders o
        JOIN CafeTables t ON t.TableID = o.TableID
        WHERE o.OrderID = :order_id
        """,
        {"order_id": order_id},
    )

    if not order:
        abort(404)

    return enrich_order(order)


def get_order_details(order_id: int) -> list[dict]:
    details = query_all(
        """
        SELECT
            od.DetailID,
            od.OrderID,
            od.ProductID,
            od.Quantity,
            od.Price,
            od.Note,
            p.ProductName,
            p.Image
        FROM OrderDetails od
        JOIN Products p ON p.ProductID = od.ProductID
        WHERE od.OrderID = :order_id
        ORDER BY od.DetailID ASC
        """,
        {"order_id": order_id},
    )

    items = []
    for row in details:
        item = dict(row)
        item["LineTotal"] = Decimal(item["Price"] or 0) * int(item["Quantity"] or 0)
        items.append(item)

    return items


def get_active_order_by_table(table_id: int):
    order = query_one(
        """
        SELECT TOP 1 o.*, t.TableName
        FROM Orders o
        JOIN CafeTables t ON t.TableID = o.TableID
        WHERE o.TableID = :table_id
          AND o.Status NOT IN (:paid_status, :cancelled_status)
        ORDER BY o.OrderID DESC
        """,
        {
            "table_id": table_id,
            "paid_status": ORDER_PAID,
            "cancelled_status": ORDER_CANCELLED,
        },
    )

    if not order:
        return None

    return enrich_order(order)


def format_order_note(sugar_level: str, ice_level: str, custom_note: str) -> str:
    parts = []

    if sugar_level:
        parts.append(f"Đường: {sugar_level}")
    if ice_level:
        parts.append(f"Đá: {ice_level}")
    if custom_note:
        parts.append(f"Ghi chú: {custom_note}")

    return " | ".join(parts)


def make_cart_key(product_id: int, sugar_level: str, ice_level: str, custom_note: str) -> str:
    raw = f"{product_id}|{sugar_level}|{ice_level}|{custom_note.strip().lower()}"
    digest = hashlib.md5(raw.encode("utf-8")).hexdigest()[:12]
    return f"{product_id}-{digest}"


def get_cart() -> list[dict]:
    return session.get("cart", [])


def save_cart(cart: list[dict]) -> None:
    session["cart"] = cart
    session.modified = True


def clear_customer_order_state() -> None:
    session.pop("cart", None)
    session.pop("voucher", None)


def build_product_lookup(product_ids: list[int]) -> dict[int, dict]:
    if not product_ids:
        return {}

    statement = text(
        """
        SELECT
            ProductID,
            ProductName,
            Price,
            Image,
            Description,
            CategoryID,
            Status
        FROM Products
        WHERE ProductID IN :product_ids
        """
    ).bindparams(bindparam("product_ids", expanding=True))

    with db.engine.connect() as connection:
        rows = connection.execute(statement, {"product_ids": product_ids}).mappings().all()

    return {row["ProductID"]: dict(row) for row in rows}


def cart_summary(cart: list[dict]) -> dict:
    product_lookup = build_product_lookup([item["product_id"] for item in cart])
    items = []
    subtotal = Decimal("0")

    for cart_item in cart:
        product = product_lookup.get(cart_item["product_id"])
        if not product:
            continue

        unit_price = Decimal(product["Price"] or 0)
        quantity = int(cart_item["quantity"])
        line_total = unit_price * quantity
        subtotal += line_total

        items.append(
            {
                "key": cart_item["key"],
                "product_id": product["ProductID"],
                "name": product["ProductName"],
                "price": unit_price,
                "quantity": quantity,
                "image": product["Image"],
                "description": product.get("Description"),
                "note": cart_item.get("note", ""),
                "sugar_level": cart_item.get("sugar_level", ""),
                "ice_level": cart_item.get("ice_level", ""),
                "line_total": line_total,
            }
        )

    voucher = session.get("voucher", "").upper()
    discount_value = discount_amount_from_code(voucher, subtotal)
    final_total = max(subtotal - discount_value, Decimal("0"))

    return {
        "items": items,
        "subtotal": subtotal,
        "discount_code": voucher,
        "discount_value": discount_value,
        "final_total": final_total,
    }


def discount_amount_from_code(code: str, subtotal: Decimal) -> Decimal:
    if not code:
        return Decimal("0")

    ratio = DISCOUNT_CODES.get(code.upper())
    if ratio is None:
        return Decimal("0")

    return (subtotal * ratio).quantize(Decimal("1.00"))


def notification_label(kind: str, table_name: str) -> str:
    labels = {
        NOTIFY_ORDER: f"Đơn mới từ {table_name}",
        NOTIFY_SUPPORT: f"{table_name} cần hỗ trợ",
        NOTIFY_PAYMENT: f"{table_name} yêu cầu thanh toán",
    }
    return labels.get(kind, f"Thông báo từ {table_name}")


def encode_notification_message(kind: str, label: str) -> str:
    return f"{kind}|{label}"


def decode_notification_message(message: str | None) -> tuple[str, str]:
    if not message or "|" not in message:
        return "INFO", message or "Thông báo"

    kind, label = message.split("|", 1)
    return kind, label


def create_notification(connection, table_id: int, kind: str) -> None:
    table = connection.execute(
        text(
            """
            SELECT TableName
            FROM CafeTables
            WHERE TableID = :table_id
            """
        ),
        {"table_id": table_id},
    ).mappings().first()

    if not table:
        return

    label = notification_label(kind, table["TableName"])
    message = encode_notification_message(kind, label)

    existing = connection.execute(
        text(
            """
            SELECT TOP 1 NotifyID
            FROM Notifications
            WHERE TableID = :table_id
              AND Status = 0
              AND Message = :message
            ORDER BY NotifyID DESC
            """
        ),
        {"table_id": table_id, "message": message},
    ).fetchone()

    if existing:
        return

    connection.execute(
        text(
            """
            INSERT INTO Notifications(TableID, Message, Status)
            VALUES (:table_id, :message, 0)
            """
        ),
        {"table_id": table_id, "message": message},
    )


def resolve_notifications(connection, table_id: int, kinds: list[str] | None = None) -> None:
    if not kinds:
        connection.execute(
            text(
                """
                UPDATE Notifications
                SET Status = 1
                WHERE TableID = :table_id AND Status = 0
                """
            ),
            {"table_id": table_id},
        )
        return

    conditions = " OR ".join(f"Message LIKE :kind_{index}" for index, _ in enumerate(kinds))
    params = {"table_id": table_id}
    for index, kind in enumerate(kinds):
        params[f"kind_{index}"] = f"{kind}|%"

    connection.execute(
        text(
            f"""
            UPDATE Notifications
            SET Status = 1
            WHERE TableID = :table_id
              AND Status = 0
              AND ({conditions})
            """
        ),
        params,
    )


def recalculate_order_total(connection, order_id: int) -> Decimal:
    total = connection.execute(
        text(
            """
            SELECT SUM(Quantity * Price)
            FROM OrderDetails
            WHERE OrderID = :order_id
            """
        ),
        {"order_id": order_id},
    ).scalar()

    amount = Decimal(total or 0)

    connection.execute(
        text(
            """
            UPDATE Orders
            SET TotalAmount = :amount
            WHERE OrderID = :order_id
            """
        ),
        {"amount": amount, "order_id": order_id},
    )

    return amount


def update_table_status(connection, table_id: int, status: str) -> None:
    connection.execute(
        text(
            """
            UPDATE CafeTables
            SET Status = :status
            WHERE TableID = :table_id
            """
        ),
        {"status": status, "table_id": table_id},
    )


def report_filters_from_request() -> dict:
    today = date.today()
    default_start = today - timedelta(days=6)

    start_raw = request.args.get("start_date", default_start.isoformat())
    end_raw = request.args.get("end_date", today.isoformat())
    group_by = request.args.get("group_by", "day").lower()

    try:
        start_date = date.fromisoformat(start_raw)
        end_date = date.fromisoformat(end_raw)
    except ValueError:
        flash("Khoảng thời gian không hợp lệ, hệ thống đã dùng mặc định 7 ngày gần nhất.", "error")
        start_date = default_start
        end_date = today

    if start_date > end_date:
        flash("Ngày bắt đầu không được lớn hơn ngày kết thúc.", "error")
        start_date = default_start
        end_date = today

    if group_by not in {"day", "month", "year"}:
        group_by = "day"

    return {
        "start_date": start_date,
        "end_date": end_date,
        "group_by": group_by,
    }


def report_group_sql(group_by: str) -> tuple[str, str]:
    if group_by == "month":
        return (
            "FORMAT(p.PaymentDate, 'MM/yyyy')",
            "DATEFROMPARTS(YEAR(p.PaymentDate), MONTH(p.PaymentDate), 1)",
        )

    if group_by == "year":
        return (
            "CAST(YEAR(p.PaymentDate) AS VARCHAR(4))",
            "DATEFROMPARTS(YEAR(p.PaymentDate), 1, 1)",
        )

    return (
        "FORMAT(p.PaymentDate, 'dd/MM/yyyy')",
        "CAST(p.PaymentDate AS DATE)",
    )


def build_report_data(start_date: date, end_date: date, group_by: str) -> dict:
    summary = query_one(
        """
        SELECT
            COUNT(*) AS PaymentCount,
            COUNT(DISTINCT p.OrderID) AS PaidOrders,
            ISNULL(SUM(p.Amount), 0) AS Revenue
        FROM Payments p
        WHERE CAST(p.PaymentDate AS DATE) BETWEEN :start_date AND :end_date
        """,
        {"start_date": start_date, "end_date": end_date},
    ) or {"PaymentCount": 0, "PaidOrders": 0, "Revenue": 0}

    average_bill = Decimal(summary["Revenue"] or 0)
    payment_count = int(summary["PaymentCount"] or 0)
    if payment_count:
        average_bill = (average_bill / Decimal(payment_count)).quantize(Decimal("1.00"))

    label_sql, order_sql = report_group_sql(group_by)
    chart_rows = query_all(
        f"""
        SELECT
            {label_sql} AS PeriodLabel,
            ISNULL(SUM(p.Amount), 0) AS Revenue,
            {order_sql} AS SortKey
        FROM Payments p
        WHERE CAST(p.PaymentDate AS DATE) BETWEEN :start_date AND :end_date
        GROUP BY {label_sql}, {order_sql}
        ORDER BY SortKey ASC
        """,
        {"start_date": start_date, "end_date": end_date},
    )

    chart_data = []
    max_revenue = max([Decimal(row["Revenue"] or 0) for row in chart_rows], default=Decimal("0"))

    for row in chart_rows:
        revenue = Decimal(row["Revenue"] or 0)
        width = 0
        if max_revenue > 0:
            width = int((revenue / max_revenue) * 100)
        chart_data.append(
            {
                "label": row["PeriodLabel"],
                "revenue": revenue,
                "width": max(width, 8 if revenue > 0 else 0),
            }
        )

    top_products = query_all(
        """
        SELECT TOP 5
            p.ProductName,
            SUM(od.Quantity) AS Sold,
            SUM(od.Quantity * od.Price) AS Revenue
        FROM Payments pay
        JOIN Orders o ON o.OrderID = pay.OrderID
        JOIN OrderDetails od ON od.OrderID = o.OrderID
        JOIN Products p ON p.ProductID = od.ProductID
        WHERE CAST(pay.PaymentDate AS DATE) BETWEEN :start_date AND :end_date
        GROUP BY p.ProductName
        ORDER BY Sold DESC, Revenue DESC
        """,
        {"start_date": start_date, "end_date": end_date},
    )

    return {
        "summary": {
            "payment_count": payment_count,
            "paid_orders": int(summary["PaidOrders"] or 0),
            "revenue": Decimal(summary["Revenue"] or 0),
            "average_bill": average_bill,
        },
        "chart_data": chart_data,
        "top_products": top_products,
    }


@app.context_processor
def inject_layout_context():
    return {
        "nav_links": make_nav_links(session.get("role")),
        "today_value": date.today(),
    }


@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        user = query_one(
            """
            SELECT UserID, Username, Password, FullName, Role, Status
            FROM Users
            WHERE Username = :username
            """,
            {"username": username},
        )

        if not user or not check_password(user["Password"], password):
            flash("Tên đăng nhập hoặc mật khẩu không đúng.", "error")
            return render_template("login.html")

        if user["Status"] in (0, False):
            flash("Tài khoản đang bị khóa.", "error")
            return render_template("login.html")

        session["user_id"] = user["UserID"]
        session["role"] = user["Role"]
        session["full_name"] = user["FullName"]

        if user["Role"] == ROLE_ADMIN:
            return redirect(url_for("admin_dashboard"))

        return redirect(url_for("staff_dashboard"))

    if session.get("role") == ROLE_ADMIN:
        return redirect(url_for("admin_dashboard"))

    if session.get("role") == ROLE_STAFF:
        return redirect(url_for("staff_dashboard"))

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("Bạn đã đăng xuất khỏi hệ thống.", "success")
    return redirect(url_for("login"))


@app.route("/admin")
@require_role(ROLE_ADMIN)
def admin_dashboard():
    dashboard = query_one(
        """
        SELECT
            (SELECT COUNT(*) FROM Products) AS ProductCount,
            (SELECT COUNT(*) FROM Users WHERE Role = :staff_role) AS StaffCount,
            (SELECT COUNT(*)
             FROM Orders
             WHERE Status IN (
                :pending_status,
                :processing_status,
                :legacy_pending_status,
                :legacy_processing_status
             )) AS ActiveOrders,
            (SELECT ISNULL(SUM(Amount), 0)
             FROM Payments
             WHERE CAST(PaymentDate AS DATE) = CAST(GETDATE() AS DATE)) AS TodayRevenue
        """,
        {
            "staff_role": ROLE_STAFF,
            "pending_status": ORDER_PENDING,
            "processing_status": ORDER_PROCESSING,
            "legacy_pending_status": "Đang chờ",
            "legacy_processing_status": "Đang pha chế",
        },
    )

    recent_products = query_all(
        """
        SELECT TOP 5
            p.ProductID,
            p.ProductName,
            p.Price,
            p.Image,
            p.Status,
            c.CategoryName
        FROM Products p
        LEFT JOIN Categories c ON c.CategoryID = p.CategoryID
        ORDER BY p.ProductID DESC
        """
    )

    return render_template(
        "admin_dashboard.html",
        stats=dashboard,
        recent_products=recent_products,
    )


@app.route("/admin/products", methods=["GET", "POST"])
@require_role(ROLE_ADMIN)
def admin_products():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        price = request.form.get("price", "").strip()
        category_id = request.form.get("category_id", "").strip()
        image = request.form.get("image", "").strip()
        description = request.form.get("description", "").strip()
        status = 1 if request.form.get("is_available") == "on" else 0

        if not name or not price or not category_id:
            flash("Vui lòng nhập đầy đủ tên món, giá và danh mục.", "error")
        else:
            with db.engine.begin() as connection:
                connection.execute(
                    text(
                        """
                        INSERT INTO Products
                        (ProductName, Price, CategoryID, Image, Description, Status)
                        VALUES (:name, :price, :category_id, :image, :description, :status)
                        """
                    ),
                    {
                        "name": name,
                        "price": Decimal(price),
                        "category_id": int(category_id),
                        "image": image or None,
                        "description": description or None,
                        "status": status,
                    },
                )

            flash("Đã thêm món mới vào thực đơn.", "success")
            return redirect(url_for("admin_products"))

    products = query_all(
        """
        SELECT
            p.ProductID,
            p.ProductName,
            p.Price,
            p.Image,
            p.Description,
            p.Status,
            c.CategoryName
        FROM Products p
        LEFT JOIN Categories c ON c.CategoryID = p.CategoryID
        ORDER BY p.ProductID DESC
        """
    )

    categories = query_all(
        """
        SELECT CategoryID, CategoryName
        FROM Categories
        ORDER BY CategoryName ASC
        """
    )

    return render_template(
        "admin_products.html",
        products=products,
        categories=categories,
        image_options=get_product_images(),
    )


@app.route("/admin/products/<int:product_id>/edit", methods=["GET", "POST"])
@require_role(ROLE_ADMIN)
def edit_product(product_id: int):
    product = query_one(
        """
        SELECT *
        FROM Products
        WHERE ProductID = :product_id
        """,
        {"product_id": product_id},
    )

    if not product:
        abort(404)

    if request.method == "POST":
        with db.engine.begin() as connection:
            connection.execute(
                text(
                    """
                    UPDATE Products
                    SET ProductName = :name,
                        Price = :price,
                        CategoryID = :category_id,
                        Image = :image,
                        Description = :description,
                        Status = :status
                    WHERE ProductID = :product_id
                    """
                ),
                {
                    "name": request.form.get("name", "").strip(),
                    "price": Decimal(request.form.get("price", "0")),
                    "category_id": int(request.form.get("category_id")),
                    "image": request.form.get("image", "").strip() or None,
                    "description": request.form.get("description", "").strip() or None,
                    "status": 1 if request.form.get("is_available") == "on" else 0,
                    "product_id": product_id,
                },
            )

        flash("Thông tin món đã được cập nhật.", "success")
        return redirect(url_for("admin_products"))

    categories = query_all(
        """
        SELECT CategoryID, CategoryName
        FROM Categories
        ORDER BY CategoryName ASC
        """
    )

    return render_template(
        "edit_product.html",
        product=product,
        categories=categories,
        image_options=get_product_images(),
    )


@app.route("/admin/products/<int:product_id>/toggle", methods=["POST"])
@require_role(ROLE_ADMIN)
def toggle_product_status(product_id: int):
    product = query_one(
        """
        SELECT ProductID, Status
        FROM Products
        WHERE ProductID = :product_id
        """,
        {"product_id": product_id},
    )

    if not product:
        abort(404)

    new_status = 0 if product["Status"] else 1

    with db.engine.begin() as connection:
        connection.execute(
            text(
                """
                UPDATE Products
                SET Status = :status
                WHERE ProductID = :product_id
                """
            ),
            {"status": new_status, "product_id": product_id},
        )

    flash("Trạng thái món đã được cập nhật.", "success")
    return redirect(url_for("admin_products"))


@app.route("/admin/users", methods=["GET", "POST"])
@require_role(ROLE_ADMIN)
def admin_users():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        full_name = request.form.get("full_name", "").strip()
        phone = request.form.get("phone", "").strip()

        if not username or not password or not full_name:
            flash("Vui lòng nhập đầy đủ tên đăng nhập, mật khẩu và họ tên.", "error")
        else:
            with db.engine.begin() as connection:
                connection.execute(
                    text(
                        """
                        INSERT INTO Users
                        (Username, Password, FullName, Role, Phone, Status)
                        VALUES (:username, :password, :full_name, :role, :phone, 1)
                        """
                    ),
                    {
                        "username": username,
                        "password": hash_password(password),
                        "full_name": full_name,
                        "role": ROLE_STAFF,
                        "phone": phone or None,
                    },
                )

            flash("Đã tạo tài khoản nhân viên mới.", "success")
            return redirect(url_for("admin_users"))

    users = query_all(
        """
        SELECT
            UserID,
            Username,
            FullName,
            Phone,
            Status
        FROM Users
        WHERE Role = :role
        ORDER BY UserID ASC
        """,
        {"role": ROLE_STAFF},
    )

    return render_template("admin_users.html", users=users)


@app.route("/admin/users/<int:user_id>/edit", methods=["GET", "POST"])
@require_role(ROLE_ADMIN)
def edit_user(user_id: int):
    user = query_one(
        """
        SELECT UserID, Username, FullName, Phone, Status
        FROM Users
        WHERE UserID = :user_id AND Role = :role
        """,
        {"user_id": user_id, "role": ROLE_STAFF},
    )

    if not user:
        abort(404)

    if request.method == "POST":
        password = request.form.get("password", "").strip()

        with db.engine.begin() as connection:
            if password:
                connection.execute(
                    text(
                        """
                        UPDATE Users
                        SET Username = :username,
                            FullName = :full_name,
                            Phone = :phone,
                            Password = :password
                        WHERE UserID = :user_id
                        """
                    ),
                    {
                        "username": request.form.get("username", "").strip(),
                        "full_name": request.form.get("full_name", "").strip(),
                        "phone": request.form.get("phone", "").strip() or None,
                        "password": hash_password(password),
                        "user_id": user_id,
                    },
                )
            else:
                connection.execute(
                    text(
                        """
                        UPDATE Users
                        SET Username = :username,
                            FullName = :full_name,
                            Phone = :phone
                        WHERE UserID = :user_id
                        """
                    ),
                    {
                        "username": request.form.get("username", "").strip(),
                        "full_name": request.form.get("full_name", "").strip(),
                        "phone": request.form.get("phone", "").strip() or None,
                        "user_id": user_id,
                    },
                )

        flash("Đã cập nhật thông tin nhân viên.", "success")
        return redirect(url_for("admin_users"))

    return render_template("edit_user.html", user=user)


@app.route("/admin/users/<int:user_id>/toggle", methods=["POST"])
@require_role(ROLE_ADMIN)
def toggle_user_status(user_id: int):
    user = query_one(
        """
        SELECT UserID, Status
        FROM Users
        WHERE UserID = :user_id AND Role = :role
        """,
        {"user_id": user_id, "role": ROLE_STAFF},
    )

    if not user:
        abort(404)

    with db.engine.begin() as connection:
        connection.execute(
            text(
                """
                UPDATE Users
                SET Status = :status
                WHERE UserID = :user_id
                """
            ),
            {"status": 0 if user["Status"] else 1, "user_id": user_id},
        )

    flash("Trạng thái tài khoản đã được cập nhật.", "success")
    return redirect(url_for("admin_users"))


@app.route("/admin/report")
@require_role(ROLE_ADMIN)
def admin_report():
    filters = report_filters_from_request()
    report_data = build_report_data(
        filters["start_date"],
        filters["end_date"],
        filters["group_by"],
    )

    return render_template(
        "admin_report.html",
        filters=filters,
        report_data=report_data,
    )


@app.route("/admin/report/excel")
@require_role(ROLE_ADMIN)
def export_report_excel():
    filters = report_filters_from_request()
    report_data = build_report_data(
        filters["start_date"],
        filters["end_date"],
        filters["group_by"],
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Bao Cao"

    title_fill = PatternFill(fill_type="solid", fgColor="4E342E")
    title_font = Font(color="FFFFFF", bold=True)

    sheet["A1"] = "BAO CAO DOANH THU QUAN CAFE"
    sheet["A2"] = f"Tu ngay {filters['start_date'].strftime('%d/%m/%Y')} den {filters['end_date'].strftime('%d/%m/%Y')}"
    sheet["A4"] = "Tong doanh thu"
    sheet["B4"] = float(report_data["summary"]["revenue"])
    sheet["A5"] = "So giao dich"
    sheet["B5"] = report_data["summary"]["payment_count"]
    sheet["A6"] = "So don da thanh toan"
    sheet["B6"] = report_data["summary"]["paid_orders"]
    sheet["A7"] = "Gia tri hoa don trung binh"
    sheet["B7"] = float(report_data["summary"]["average_bill"])

    sheet["A9"] = "Ky"
    sheet["B9"] = "Doanh thu"
    sheet["A9"].fill = title_fill
    sheet["B9"].fill = title_fill
    sheet["A9"].font = title_font
    sheet["B9"].font = title_font

    row_cursor = 10
    for point in report_data["chart_data"]:
        sheet[f"A{row_cursor}"] = point["label"]
        sheet[f"B{row_cursor}"] = float(point["revenue"])
        row_cursor += 1

    row_cursor += 1
    sheet[f"A{row_cursor}"] = "Top mon ban chay"
    row_cursor += 1
    sheet[f"A{row_cursor}"] = "Ten mon"
    sheet[f"B{row_cursor}"] = "So luong"
    sheet[f"C{row_cursor}"] = "Doanh thu"
    for column in ("A", "B", "C"):
        sheet[f"{column}{row_cursor}"].fill = title_fill
        sheet[f"{column}{row_cursor}"].font = title_font

    row_cursor += 1
    for product in report_data["top_products"]:
        sheet[f"A{row_cursor}"] = product["ProductName"]
        sheet[f"B{row_cursor}"] = int(product["Sold"] or 0)
        sheet[f"C{row_cursor}"] = float(product["Revenue"] or 0)
        row_cursor += 1

    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 18

    file_object = io.BytesIO()
    workbook.save(file_object)
    file_object.seek(0)

    return send_file(
        file_object,
        as_attachment=True,
        download_name="bao_cao_doanh_thu.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/menu/<int:table_id>")
def customer_menu(table_id: int):
    table = get_table_or_404(table_id)
    session["table_id"] = table_id

    keyword = request.args.get("keyword", "").strip()
    category_id = request.args.get("category_id", "").strip()

    sql = """
        SELECT
            p.ProductID,
            p.ProductName,
            p.Price,
            p.Image,
            p.Description,
            p.Status,
            c.CategoryName
        FROM Products p
        LEFT JOIN Categories c ON c.CategoryID = p.CategoryID
        WHERE 1 = 1
    """
    params: dict[str, str | int] = {}

    if keyword:
        sql += " AND p.ProductName LIKE :keyword"
        params["keyword"] = f"%{keyword}%"

    if category_id:
        sql += " AND p.CategoryID = :category_id"
        params["category_id"] = int(category_id)

    sql += " ORDER BY p.Status DESC, p.ProductID ASC"

    products = query_all(sql, params)
    categories = query_all(
        """
        SELECT CategoryID, CategoryName
        FROM Categories
        ORDER BY CategoryName ASC
        """
    )
    active_order = get_active_order_by_table(table_id)
    cart_info = cart_summary(get_cart())

    return render_template(
        "customer_menu.html",
        table=table,
        products=products,
        categories=categories,
        keyword=keyword,
        selected_category=category_id,
        sugar_options=SUGAR_OPTIONS,
        ice_options=ICE_OPTIONS,
        active_order=active_order,
        cart_info=cart_info,
    )


@app.route("/cart/add", methods=["POST"])
def add_to_cart():
    table_id = int(request.form.get("table_id", session.get("table_id", 0)))
    get_table_or_404(table_id)
    session["table_id"] = table_id

    product_id = int(request.form["product_id"])
    quantity = max(int(request.form.get("quantity", 1)), 1)
    sugar_level = request.form.get("sugar_level", SUGAR_OPTIONS[0]).strip()
    ice_level = request.form.get("ice_level", ICE_OPTIONS[0]).strip()
    custom_note = request.form.get("custom_note", "").strip()

    product = query_one(
        """
        SELECT ProductID, ProductName, Status
        FROM Products
        WHERE ProductID = :product_id
        """,
        {"product_id": product_id},
    )

    if not product or not product["Status"]:
        flash("Món này hiện đang tạm hết, vui lòng chọn món khác.", "error")
        return redirect(url_for("customer_menu", table_id=table_id))

    note = format_order_note(sugar_level, ice_level, custom_note)
    cart = get_cart()
    item_key = make_cart_key(product_id, sugar_level, ice_level, custom_note)

    for item in cart:
        if item["key"] == item_key:
            item["quantity"] += quantity
            save_cart(cart)
            flash("Đã cập nhật số lượng món trong giỏ.", "success")
            return redirect(url_for("customer_menu", table_id=table_id))

    cart.append(
        {
            "key": item_key,
            "product_id": product_id,
            "quantity": quantity,
            "note": note,
            "sugar_level": sugar_level,
            "ice_level": ice_level,
        }
    )
    save_cart(cart)
    flash("Đã thêm món vào giỏ hàng.", "success")
    return redirect(url_for("customer_menu", table_id=table_id))


@app.route("/cart")
def cart():
    table_id = session.get("table_id")
    if not table_id:
        flash("Vui lòng quét mã QR hoặc chọn bàn trước khi gọi món.", "error")
        return redirect(url_for("login"))

    table = get_table_or_404(int(table_id))
    summary = cart_summary(get_cart())

    return render_template(
        "customer_cart.html",
        table=table,
        cart_info=summary,
    )


@app.route("/cart/update/<item_key>/<action>")
def update_cart_item(item_key: str, action: str):
    cart = get_cart()
    updated_cart = []

    for item in cart:
        if item["key"] != item_key:
            updated_cart.append(item)
            continue

        if action == "increase":
            item["quantity"] += 1
            updated_cart.append(item)
        elif action == "decrease":
            item["quantity"] -= 1
            if item["quantity"] > 0:
                updated_cart.append(item)
        elif action == "remove":
            continue
        else:
            updated_cart.append(item)

    save_cart(updated_cart)
    flash("Giỏ hàng đã được cập nhật.", "success")
    return redirect(url_for("cart"))


@app.route("/apply-voucher", methods=["POST"])
def apply_voucher():
    code = request.form.get("voucher", "").strip().upper()
    if code in DISCOUNT_CODES:
        session["voucher"] = code
        flash("Mã giảm giá đã được áp dụng cho đơn hiện tại.", "success")
    else:
        session.pop("voucher", None)
        flash("Mã giảm giá không hợp lệ.", "error")

    return redirect(url_for("cart"))


@app.route("/checkout", methods=["POST"])
def checkout():
    table_id = session.get("table_id")
    if not table_id:
        flash("Không tìm thấy thông tin bàn.", "error")
        return redirect(url_for("login"))

    cart = get_cart()
    if not cart:
        flash("Giỏ hàng đang trống.", "error")
        return redirect(url_for("cart"))

    summary = cart_summary(cart)
    if not summary["items"]:
        flash("Không có món hợp lệ trong giỏ hàng.", "error")
        return redirect(url_for("cart"))

    with db.engine.begin() as connection:
        active_order = connection.execute(
            text(
                """
                SELECT TOP 1 OrderID
                FROM Orders
                WHERE TableID = :table_id
                  AND Status NOT IN (:paid_status, :cancelled_status)
                ORDER BY OrderID DESC
                """
            ),
            {
                "table_id": table_id,
                "paid_status": ORDER_PAID,
                "cancelled_status": ORDER_CANCELLED,
            },
        ).mappings().first()

        if active_order:
            order_id = active_order["OrderID"]
        else:
            order_id = connection.execute(
                text(
                    """
                    INSERT INTO Orders(TableID, Status, TotalAmount)
                    OUTPUT INSERTED.OrderID
                    VALUES (:table_id, :status, 0)
                    """
                ),
                {"table_id": table_id, "status": ORDER_PENDING},
            ).scalar_one()

        for item in summary["items"]:
            existing_detail = connection.execute(
                text(
                    """
                    SELECT TOP 1 DetailID, Quantity
                    FROM OrderDetails
                    WHERE OrderID = :order_id
                      AND ProductID = :product_id
                      AND ISNULL(Note, '') = :note
                    """
                ),
                {
                    "order_id": order_id,
                    "product_id": item["product_id"],
                    "note": item["note"],
                },
            ).mappings().first()

            if existing_detail:
                connection.execute(
                    text(
                        """
                        UPDATE OrderDetails
                        SET Quantity = Quantity + :quantity
                        WHERE DetailID = :detail_id
                        """
                    ),
                    {
                        "quantity": item["quantity"],
                        "detail_id": existing_detail["DetailID"],
                    },
                )
            else:
                connection.execute(
                    text(
                        """
                        INSERT INTO OrderDetails
                        (OrderID, ProductID, Quantity, Price, Note)
                        VALUES (:order_id, :product_id, :quantity, :price, :note)
                        """
                    ),
                    {
                        "order_id": order_id,
                        "product_id": item["product_id"],
                        "quantity": item["quantity"],
                        "price": item["price"],
                        "note": item["note"],
                    },
                )

        order_total = recalculate_order_total(connection, order_id)
        discount_value = summary["discount_value"]
        final_total = max(order_total - discount_value, Decimal("0"))
        connection.execute(
            text(
                """
                UPDATE Orders
                SET TotalAmount = :amount,
                    Status = :pending_status
                WHERE OrderID = :order_id
                """
            ),
            {
                "amount": final_total,
                "order_id": order_id,
                "pending_status": ORDER_PENDING,
            },
        )
        update_table_status(connection, int(table_id), TABLE_OCCUPIED)
        create_notification(connection, int(table_id), NOTIFY_ORDER)

    clear_customer_order_state()
    tracked_orders = session.get("my_orders", [])
    if order_id not in tracked_orders:
        tracked_orders.append(order_id)
        session["my_orders"] = tracked_orders

    flash("Đơn hàng đã được gửi tới quầy.", "success")
    return redirect(url_for("tracking", order_id=order_id))


@app.route("/tracking/<int:order_id>")
def tracking(order_id: int):
    order = get_order_or_404(order_id)
    items = get_order_details(order_id)
    return render_template("customer_tracking.html", order=order, items=items)


@app.route("/my-orders/<int:table_id>")
def my_orders(table_id: int):
    get_table_or_404(table_id)
    tracked_orders = session.get("my_orders", [])
    if not tracked_orders:
        orders = []
    else:
        statement = text(
            """
            SELECT o.*, t.TableName
            FROM Orders o
            JOIN CafeTables t ON t.TableID = o.TableID
            WHERE o.OrderID IN :order_ids
            ORDER BY o.OrderID DESC
            """
        ).bindparams(bindparam("order_ids", expanding=True))

        with db.engine.connect() as connection:
            orders = connection.execute(statement, {"order_ids": tracked_orders}).mappings().all()
        orders = enrich_orders(orders)

    table = get_table_or_404(table_id)
    return render_template("customer_orders.html", orders=orders, table=table)


@app.route("/customer/request", methods=["POST"])
def customer_request():
    table_id = int(request.form.get("table_id", 0))
    action = request.form.get("action", "")
    redirect_to = request.form.get("redirect_to") or url_for("customer_menu", table_id=table_id)
    get_table_or_404(table_id)

    kind = NOTIFY_SUPPORT if action == "support" else NOTIFY_PAYMENT

    with db.engine.begin() as connection:
        create_notification(connection, table_id, kind)

    if kind == NOTIFY_SUPPORT:
        flash("Yêu cầu hỗ trợ đã được gửi tới nhân viên.", "success")
    else:
        flash("Yêu cầu thanh toán đã được gửi tới quầy.", "success")

    return redirect(redirect_to)


@app.route("/staff")
@require_role(ROLE_STAFF)
def staff_dashboard():
    orders = query_all(
        """
        SELECT
            o.OrderID,
            o.TableID,
            t.TableName,
            o.OrderDate,
            o.Status,
            o.TotalAmount,
            SUM(od.Quantity) AS TotalItems
        FROM Orders o
        JOIN CafeTables t ON t.TableID = o.TableID
        LEFT JOIN OrderDetails od ON od.OrderID = o.OrderID
        WHERE o.Status NOT IN (:paid_status, :cancelled_status)
        GROUP BY o.OrderID, o.TableID, t.TableName, o.OrderDate, o.Status, o.TotalAmount
        ORDER BY o.OrderDate DESC
        """,
        {"paid_status": ORDER_PAID, "cancelled_status": ORDER_CANCELLED},
    )
    orders = enrich_orders(orders)

    tables = query_all(
        """
        SELECT
            t.TableID,
            t.TableName,
            t.Status,
            active.OrderID,
            active.Status AS OrderStatus,
            active.TotalAmount,
            pending.NotifyCount
        FROM CafeTables t
        OUTER APPLY (
            SELECT TOP 1
                o.OrderID,
                o.Status,
                o.TotalAmount
            FROM Orders o
            WHERE o.TableID = t.TableID
              AND o.Status NOT IN (:paid_status, :cancelled_status)
            ORDER BY o.OrderID DESC
        ) active
        OUTER APPLY (
            SELECT COUNT(*) AS NotifyCount
            FROM Notifications n
            WHERE n.TableID = t.TableID
              AND n.Status = 0
        ) pending
        ORDER BY t.TableID ASC
        """,
        {"paid_status": ORDER_PAID, "cancelled_status": ORDER_CANCELLED},
    )
    tables = [enrich_table(row) for row in tables]

    notifications = query_all(
        """
        SELECT
            NotifyID,
            TableID,
            Message,
            CreatedAt
        FROM Notifications
        WHERE Status = 0
        ORDER BY CreatedAt DESC, NotifyID DESC
        """
    )

    decorated_notifications = []
    for row in notifications:
        kind, label = decode_notification_message(row["Message"])
        decorated_notifications.append(
            {
                "NotifyID": row["NotifyID"],
                "TableID": row["TableID"],
                "Kind": kind,
                "Label": label,
                "CreatedAt": row["CreatedAt"],
            }
        )

    stats = {
        "pending": sum(1 for order in orders if order["Status"] == ORDER_PENDING),
        "processing": sum(1 for order in orders if order["Status"] == ORDER_PROCESSING),
        "completed": sum(1 for order in orders if order["Status"] == ORDER_COMPLETED),
        "active_tables": sum(1 for table in tables if table["Status"] == TABLE_OCCUPIED),
    }

    return render_template(
        "staff_dashboard.html",
        orders=orders,
        tables=tables,
        notifications=decorated_notifications,
        stats=stats,
    )


@app.route("/staff/order/<int:order_id>")
@require_role(ROLE_STAFF)
def staff_order_detail(order_id: int):
    order = get_order_or_404(order_id)
    items = get_order_details(order_id)
    return render_template("staff_order_detail.html", order=order, items=items)


@app.route("/staff/order/<int:order_id>/confirm", methods=["POST"])
@require_role(ROLE_STAFF)
def confirm_order(order_id: int):
    order = get_order_or_404(order_id)

    with db.engine.begin() as connection:
        connection.execute(
            text(
                """
                UPDATE Orders
                SET Status = :status
                WHERE OrderID = :order_id
                """
            ),
            {"status": ORDER_PROCESSING, "order_id": order_id},
        )
        resolve_notifications(connection, order["TableID"], [NOTIFY_ORDER])

    updated_order = get_order_or_404(order_id)
    items = get_order_details(order_id)
    flash("Đơn hàng đã được xác nhận và chuyển sang pha chế.", "success")
    return render_template("staff_ticket.html", order=updated_order, items=items, auto_print=True)


@app.route("/staff/order/<int:order_id>/complete", methods=["POST"])
@require_role(ROLE_STAFF)
def complete_order(order_id: int):
    order = get_order_or_404(order_id)

    if order["Status"] == ORDER_PAID:
        flash("Đơn này đã được thanh toán.", "error")
        return redirect(url_for("staff_order_detail", order_id=order_id))

    with db.engine.begin() as connection:
        connection.execute(
            text(
                """
                UPDATE Orders
                SET Status = :status
                WHERE OrderID = :order_id
                """
            ),
            {"status": ORDER_COMPLETED, "order_id": order_id},
        )

    flash("Đơn hàng đã được cập nhật thành hoàn thành.", "success")
    return redirect(url_for("staff_order_detail", order_id=order_id))


@app.route("/staff/order/<int:order_id>/edit", methods=["GET", "POST"])
@require_role(ROLE_STAFF)
def edit_order(order_id: int):
    order = get_order_or_404(order_id)
    if order["Status"] in {ORDER_COMPLETED, ORDER_PAID}:
        flash("Không thể sửa hoặc hủy đơn đã hoàn thành/thanh toán.", "error")
        return redirect(url_for("staff_order_detail", order_id=order_id))

    items = get_order_details(order_id)

    if request.method == "POST":
        action = request.form.get("action", "update")

        with db.engine.begin() as connection:
            if action == "cancel":
                connection.execute(
                    text(
                        """
                        DELETE FROM OrderDetails
                        WHERE OrderID = :order_id
                        """
                    ),
                    {"order_id": order_id},
                )
                connection.execute(
                    text(
                        """
                        UPDATE Orders
                        SET Status = :status,
                            TotalAmount = 0
                        WHERE OrderID = :order_id
                        """
                    ),
                    {"status": ORDER_CANCELLED, "order_id": order_id},
                )
                update_table_status(connection, order["TableID"], TABLE_EMPTY)
                resolve_notifications(connection, order["TableID"])
                flash("Đơn hàng đã được hủy.", "success")
                return redirect(url_for("staff_dashboard"))

            for item in items:
                quantity_raw = request.form.get(f"quantity_{item['DetailID']}", str(item["Quantity"]))
                note_raw = request.form.get(f"note_{item['DetailID']}", item.get("Note") or "").strip()
                quantity = max(int(quantity_raw or 0), 0)

                if quantity == 0:
                    connection.execute(
                        text(
                            """
                            DELETE FROM OrderDetails
                            WHERE DetailID = :detail_id
                            """
                        ),
                        {"detail_id": item["DetailID"]},
                    )
                else:
                    connection.execute(
                        text(
                            """
                            UPDATE OrderDetails
                            SET Quantity = :quantity,
                                Note = :note
                            WHERE DetailID = :detail_id
                            """
                        ),
                        {
                            "quantity": quantity,
                            "note": note_raw or None,
                            "detail_id": item["DetailID"],
                        },
                    )

            new_total = recalculate_order_total(connection, order_id)
            if new_total <= 0:
                connection.execute(
                    text(
                        """
                        UPDATE Orders
                        SET Status = :status
                        WHERE OrderID = :order_id
                        """
                    ),
                    {"status": ORDER_CANCELLED, "order_id": order_id},
                )
                update_table_status(connection, order["TableID"], TABLE_EMPTY)
                resolve_notifications(connection, order["TableID"])
                flash("Đơn hàng đã được hủy vì không còn món nào.", "success")
                return redirect(url_for("staff_dashboard"))

        flash("Chi tiết đơn hàng đã được cập nhật.", "success")
        return redirect(url_for("staff_order_detail", order_id=order_id))

    return render_template("staff_order_edit.html", order=order, items=items)


@app.route("/staff/notifications/<int:notify_id>/resolve", methods=["POST"])
@require_role(ROLE_STAFF)
def resolve_notification(notify_id: int):
    with db.engine.begin() as connection:
        notification = connection.execute(
            text(
                """
                SELECT NotifyID, TableID, Message
                FROM Notifications
                WHERE NotifyID = :notify_id
                """
            ),
            {"notify_id": notify_id},
        ).mappings().first()

        if not notification:
            abort(404)

        connection.execute(
            text(
                """
                UPDATE Notifications
                SET Status = 1
                WHERE NotifyID = :notify_id
                """
            ),
            {"notify_id": notify_id},
        )

    kind, _ = decode_notification_message(notification["Message"])
    if kind == NOTIFY_PAYMENT:
        return redirect(url_for("payment_page", table_id=notification["TableID"]))

    flash("Thông báo đã được xử lý.", "success")
    return redirect(url_for("staff_dashboard"))


@app.route("/staff/payment/<int:table_id>", methods=["GET", "POST"])
@require_role(ROLE_STAFF)
def payment_page(table_id: int):
    table = get_table_or_404(table_id)
    order = get_active_order_by_table(table_id)

    if not order:
        flash("Bàn này hiện không có đơn chờ thanh toán.", "error")
        return redirect(url_for("staff_dashboard"))

    items = get_order_details(order["OrderID"])
    subtotal = sum((item["LineTotal"] for item in items), start=Decimal("0"))
    discount_code = request.form.get("discount_code", "").strip().upper() if request.method == "POST" else ""
    payment_method = request.form.get("payment_method", "Tiền mặt")
    discount_value = discount_amount_from_code(discount_code, subtotal)
    final_total = max(subtotal - discount_value, Decimal("0"))

    if request.method == "POST" and request.form.get("action") == "confirm":
        with db.engine.begin() as connection:
            connection.execute(
                text(
                    """
                    UPDATE Orders
                    SET Status = :status,
                        TotalAmount = :amount
                    WHERE OrderID = :order_id
                    """
                ),
                {
                    "status": ORDER_PAID,
                    "amount": final_total,
                    "order_id": order["OrderID"],
                },
            )
            payment_id = connection.execute(
                text(
                    """
                    INSERT INTO Payments(OrderID, Method, Amount)
                    OUTPUT INSERTED.PaymentID
                    VALUES (:order_id, :method, :amount)
                    """
                ),
                {
                    "order_id": order["OrderID"],
                    "method": payment_method,
                    "amount": final_total,
                },
            ).scalar_one()
            update_table_status(connection, table_id, TABLE_EMPTY)
            resolve_notifications(connection, table_id)

        paid_order = get_order_or_404(order["OrderID"])
        payment = query_one(
            """
            SELECT PaymentID, PaymentDate, Method, Amount
            FROM Payments
            WHERE PaymentID = :payment_id
            """,
            {"payment_id": payment_id},
        )

        flash("Thanh toán thành công.", "success")
        return render_template(
            "staff_bill.html",
            order=paid_order,
            table=table,
            items=items,
            payment=payment,
            subtotal=subtotal,
            discount_code=discount_code,
            discount_value=discount_value,
            final_total=final_total,
            auto_print=True,
        )

    return render_template(
        "staff_payment.html",
        order=order,
        table=table,
        items=items,
        subtotal=subtotal,
        discount_code=discount_code,
        discount_value=discount_value,
        final_total=final_total,
        payment_method=payment_method,
        valid_discount=bool(discount_code and discount_code in DISCOUNT_CODES),
    )


@app.route("/staff/bill/<int:order_id>")
@require_role(ROLE_STAFF)
def staff_bill(order_id: int):
    order = get_order_or_404(order_id)
    items = get_order_details(order_id)
    payment = query_one(
        """
        SELECT TOP 1 PaymentID, PaymentDate, Method, Amount
        FROM Payments
        WHERE OrderID = :order_id
        ORDER BY PaymentID DESC
        """,
        {"order_id": order_id},
    )
    subtotal = sum((item["LineTotal"] for item in items), start=Decimal("0"))

    return render_template(
        "staff_bill.html",
        order=order,
        table=get_table_or_404(order["TableID"]),
        items=items,
        payment=payment,
        subtotal=subtotal,
        discount_code="",
        discount_value=Decimal("0"),
        final_total=Decimal(payment["Amount"] if payment else order["TotalAmount"] or 0),
        auto_print=False,
    )


if __name__ == "__main__":
    app.run(debug=True)
